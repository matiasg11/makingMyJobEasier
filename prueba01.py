import openpyxl

wb = openpyxl.load_workbook("bajada_lastres.xlsx")
# print(type(wb))

ppal = wb.active
wb.create_sheet("Certificado")
ppal.title = "Boca"
for x in range(6, 1000):
    if ppal.cell(row=x, column=3).value != None:
        wb["Certificado"].cell(row=x, column=1).value = (
            wb["Boca"].cell(row=x, column=3).value
        )
    else:
        break


wb.save("bajada_sap2.xlsx")
