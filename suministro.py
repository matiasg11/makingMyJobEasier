import openpyxl
import datetime
import math

# Crear workbook y guardarlo con el nombre
workbook = openpyxl.Workbook()
workbook.save("suministro.xlsx")

# Crear hoja y nombrarla
workbook.create_sheet()
sheet = workbook.active
sheet.title = "Arena"

hoy = datetime.date.today()


def next_weekday(d, weekday):
    days_ahead = weekday - d.weekday()
    if days_ahead < 0:  # Target day already happened this week
        days_ahead += 7
    return d + datetime.timedelta(days_ahead)


d = datetime.date.today()
next_sunday = next_weekday(d, 7)
dates = [next_weekday(next_sunday, i) for i in range(0, 7)]

carga = 3200
cascotes = 0.4

arena_semanal = (2.1 * (carga / 7) * (1 - cascotes) / 2.877) * 7
arena_silos = [math.ceil(arena_semanal / 60) * 10 for dia in range(0, 6)]
arena_silos.append(0)
camiones_silos = [math.ceil(arena_semanal / (6 * 32)) for dia in range(0, 6)]
camiones_silos.append(0)
arena_recuperacion = 600
arena_piletas = [math.ceil(arena_recuperacion / 60) * 10 for dia in range(0, 6)]
arena_piletas.append(0)
camiones_piletas = [math.ceil(arena_recuperacion / (6 * 32)) for dia in range(0, 6)]
camiones_piletas.append(0)
arena_total = [arena_piletas[i] + arena_silos[i] for i in range(0, 7)]
camiones_total = [camiones_piletas[i] + camiones_silos[i] for i in range(0, 7)]
filas = range(1, 8)

for i, date, a_silo, c_silo, a_pile, c_pile, a_tot, c_tot in zip(
    filas,
    dates,
    arena_silos,
    camiones_silos,
    arena_piletas,
    camiones_piletas,
    arena_total,
    camiones_total,
):
    sheet.cell(row=i, column=2, value=date)
    sheet.cell(row=i, column=3, value=a_silo)
    sheet.cell(row=i, column=4, value=c_silo)
    sheet.cell(row=i, column=5, value=a_pile)
    sheet.cell(row=i, column=6, value=c_pile)
    sheet.cell(row=i, column=7, value=a_tot)
    sheet.cell(row=i, column=8, value=c_tot)


workbook.save("suministro.xlsx")
