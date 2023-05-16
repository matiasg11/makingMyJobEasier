import openpyxl

wb = openpyxl.load_workbook("bajada_lastres.xlsx")
# print(type(wb))

ppal = wb.active
wb.create_sheet("Certificado")
ppal.title = "Boca"
material = wb["Boca"].cell(row=5, column=2).value
start = 1

for row in range(6, 1000):
    if ppal.cell(row=row, column=3).value == None:
        material = wb["Boca"].cell(row=row, column=2).value
        start += 1
    elif ppal.cell(row=row, column=3).value < 200:
        wb["Certificado"].cell(row=start, column=1).value = material
        wb["Certificado"].cell(row=start, column=2).value = (
            wb["Boca"].cell(row=row, column=18).value
        )
        wb["Certificado"].cell(row=start, column=3).value = (
            wb["Boca"].cell(row=row, column=19).value
        )
        wb["Certificado"].cell(row=start, column=4).value = (
            wb["Boca"].cell(row=row, column=20).value
        )
        wb["Certificado"].cell(row=start, column=5).value = (
            wb["Boca"].cell(row=row, column=21).value
        )
        wb["Certificado"].cell(row=start, column=6).value = (
            wb["Boca"].cell(row=row, column=22).value
        )
        start += 1
    else:
        wb["Certificado"].delete_rows(start)


wb.save("bajada_sap2.xlsx")
